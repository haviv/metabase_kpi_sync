#!/usr/bin/env python3
"""
GitHub Actions Test Results Extractor

Fetches test results from GitHub Actions workflow runs, parses various report formats,
and stores results in PostgreSQL for Metabase reporting.

Supports multiple repositories and report formats:
- TRX (Visual Studio Test Results XML) - .NET tests
- Extent Summary (test-summary.txt) - Java/TestNG tests
- Postman HTML (Newman reports) - API tests

Configuration (via .env):
    GITHUB_TOKEN: GitHub Personal Access Token with 'actions:read' permission
    GITHUB_TEST_CONFIGS: JSON array of repository configurations
    
    Each config object:
    {
        "owner": "Pathlock",
        "repo": "pathlock-plc",
        "branch": "cloud/dev",  # null for no branch filter
        "workflows": ["Workflow Name 1", "Workflow Name 2"],
        "artifact_patterns": ["test-results", "nightly-test-results"],
        "report_format": "trx",  # trx, extent_summary, or postman_html
        "team": "Platform"
    }
"""

import os
import io
import re
import json
import zipfile
import requests
import xml.etree.ElementTree as ET
from datetime import datetime, timedelta
from pathlib import Path
from dotenv import load_dotenv
from sqlalchemy import create_engine, text, Table, Column, Integer, BigInteger, String, DateTime, Float, MetaData, inspect
from sqlalchemy.dialects.postgresql import TEXT as PG_TEXT
from sqlalchemy.exc import SQLAlchemyError

# Load environment variables
env_path = Path(__file__).parent / '.env'
load_dotenv(dotenv_path=env_path)

# TRX XML namespace
TRX_NAMESPACE = {'trx': 'http://microsoft.com/schemas/VisualStudio/TeamTest/2010'}


class TRXParser:
    """Parser for Visual Studio Test Results XML (TRX) files."""
    
    @staticmethod
    def parse(trx_content: str) -> dict:
        """
        Parse TRX XML content and extract test results.
        
        Args:
            trx_content: String content of the TRX file
            
        Returns:
            Dictionary containing:
            - run_info: Test run metadata
            - summary: Pass/fail counts
            - tests: List of individual test results
        """
        try:
            # Remove BOM if present
            if trx_content.startswith('\ufeff'):
                trx_content = trx_content[1:]
            
            root = ET.fromstring(trx_content)
            
            # Extract run info
            run_info = {
                'id': root.get('id', ''),
                'name': root.get('name', ''),
                'run_user': root.get('runUser', ''),
            }
            
            # Extract times
            times_elem = root.find('trx:Times', TRX_NAMESPACE)
            if times_elem is not None:
                run_info['creation_time'] = TRXParser._parse_datetime(times_elem.get('creation'))
                run_info['start_time'] = TRXParser._parse_datetime(times_elem.get('start'))
                run_info['finish_time'] = TRXParser._parse_datetime(times_elem.get('finish'))
            
            # Extract test settings for computer name
            test_settings = root.find('trx:TestSettings', TRX_NAMESPACE)
            if test_settings is not None:
                deployment = test_settings.find('trx:Deployment', TRX_NAMESPACE)
                if deployment is not None:
                    run_info['deployment_root'] = deployment.get('runDeploymentRoot', '')
            
            # Build a map of test ID -> class name from TestDefinitions
            test_class_map = TRXParser._build_test_class_map(root)
            
            # Extract individual test results
            tests = []
            results_elem = root.find('trx:Results', TRX_NAMESPACE)
            
            if results_elem is not None:
                for result in results_elem.findall('trx:UnitTestResult', TRX_NAMESPACE):
                    test_data = TRXParser._parse_test_result(result)
                    # Look up the class name from test definitions
                    test_id = test_data.get('test_id', '')
                    if test_id and test_id in test_class_map:
                        test_data['test_class'] = test_class_map[test_id]
                    tests.append(test_data)
            
            # Calculate summary
            summary = {
                'total': len(tests),
                'passed': sum(1 for t in tests if t['outcome'] == 'Passed'),
                'failed': sum(1 for t in tests if t['outcome'] == 'Failed'),
                'not_executed': sum(1 for t in tests if t['outcome'] == 'NotExecuted'),
            }
            
            # Calculate total duration
            total_duration_ms = sum(t['duration_ms'] or 0 for t in tests)
            run_info['total_duration_seconds'] = total_duration_ms / 1000
            
            return {
                'run_info': run_info,
                'summary': summary,
                'tests': tests
            }
            
        except ET.ParseError as e:
            print(f"Error parsing TRX XML: {e}")
            return None
    
    @staticmethod
    def _build_test_class_map(root) -> dict:
        """
        Build a map of test ID -> class name from TestDefinitions section.
        
        The TRX structure has:
        <TestDefinitions>
          <UnitTest id="...">
            <TestMethod className="Namespace.ClassName" name="MethodName" />
          </UnitTest>
        </TestDefinitions>
        """
        test_class_map = {}
        
        test_definitions = root.find('trx:TestDefinitions', TRX_NAMESPACE)
        if test_definitions is not None:
            for unit_test in test_definitions.findall('trx:UnitTest', TRX_NAMESPACE):
                test_id = unit_test.get('id', '')
                if not test_id:
                    continue
                
                # Find TestMethod element which contains the className
                test_method = unit_test.find('trx:TestMethod', TRX_NAMESPACE)
                if test_method is not None:
                    class_name = test_method.get('className', '')
                    if class_name:
                        test_class_map[test_id] = class_name
        
        return test_class_map
    
    @staticmethod
    def _parse_test_result(result_elem) -> dict:
        """Parse a single UnitTestResult element."""
        test_data = {
            'execution_id': result_elem.get('executionId', ''),
            'test_id': result_elem.get('testId', ''),
            'test_name': result_elem.get('testName', ''),
            'computer_name': result_elem.get('computerName', ''),
            'outcome': result_elem.get('outcome', ''),
            'start_time': TRXParser._parse_datetime(result_elem.get('startTime')),
            'end_time': TRXParser._parse_datetime(result_elem.get('endTime')),
            'duration_ms': TRXParser._parse_duration(result_elem.get('duration')),
            'error_message': None,
            'stack_trace': None,
            'stdout': None,
            'test_class': None,  # Will be populated from TestDefinitions
        }
        
        # Extract output information
        output_elem = result_elem.find('trx:Output', TRX_NAMESPACE)
        if output_elem is not None:
            # Standard output
            stdout_elem = output_elem.find('trx:StdOut', TRX_NAMESPACE)
            if stdout_elem is not None and stdout_elem.text:
                test_data['stdout'] = stdout_elem.text[:5000]  # Truncate to 5000 chars
            
            # Error info
            error_info = output_elem.find('trx:ErrorInfo', TRX_NAMESPACE)
            if error_info is not None:
                message_elem = error_info.find('trx:Message', TRX_NAMESPACE)
                if message_elem is not None and message_elem.text:
                    test_data['error_message'] = message_elem.text[:2000]  # Truncate
                
                stack_elem = error_info.find('trx:StackTrace', TRX_NAMESPACE)
                if stack_elem is not None and stack_elem.text:
                    test_data['stack_trace'] = stack_elem.text[:5000]  # Truncate
        
        return test_data
    
    @staticmethod
    def _parse_datetime(dt_str: str) -> datetime:
        """Parse datetime string from TRX format."""
        if not dt_str:
            return None
        
        # Try various formats
        formats = [
            "%Y-%m-%dT%H:%M:%S.%f%z",
            "%Y-%m-%dT%H:%M:%S.%fZ",
            "%Y-%m-%dT%H:%M:%S%z",
            "%Y-%m-%dT%H:%M:%SZ",
        ]
        
        for fmt in formats:
            try:
                # Handle timezone offset format (+00:00)
                if '+' in dt_str and ':' in dt_str.split('+')[-1]:
                    # Remove colon from timezone offset for parsing
                    dt_str_clean = dt_str[:-3] + dt_str[-2:]
                    return datetime.strptime(dt_str_clean, fmt)
                return datetime.strptime(dt_str, fmt)
            except ValueError:
                continue
        
        # Try without timezone
        try:
            # Strip timezone info and parse
            dt_clean = re.sub(r'[+-]\d{2}:\d{2}$', '', dt_str)
            return datetime.strptime(dt_clean, "%Y-%m-%dT%H:%M:%S.%f")
        except ValueError:
            pass
        
        return None
    
    @staticmethod
    def _parse_duration(duration_str: str) -> float:
        """
        Parse duration string from TRX format (HH:MM:SS.FFFFFFF) to milliseconds.
        """
        if not duration_str:
            return None
        
        try:
            parts = duration_str.split(':')
            if len(parts) == 3:
                hours = int(parts[0])
                minutes = int(parts[1])
                seconds = float(parts[2])
                return (hours * 3600 + minutes * 60 + seconds) * 1000
        except (ValueError, IndexError):
            pass
        
        return None


class ExtentSummaryParser:
    """Parser for Extent Reports test-summary.txt files (Java/TestNG tests)."""
    
    @staticmethod
    def parse(content: str) -> dict:
        """
        Parse test-summary.txt content.
        
        Expected format:
        Total Passed  : 139
        Total Failed  : 2
        Total Skipped : 0
        Total Tests   : 141
        Total Duration: 03:19:25
        
        Returns:
            Dictionary containing run_info, summary, and empty tests list
        """
        try:
            result = {
                'run_info': {
                    'id': '',
                    'name': 'Extent Report Summary',
                    'total_duration_seconds': 0,
                },
                'summary': {
                    'total': 0,
                    'passed': 0,
                    'failed': 0,
                    'not_executed': 0,
                },
                'tests': []  # No individual test details from summary
            }
            
            # Parse each line
            for line in content.strip().split('\n'):
                line = line.strip()
                if ':' not in line:
                    continue
                    
                key, value = line.split(':', 1)
                key = key.strip().lower()
                value = value.strip()
                
                if 'passed' in key:
                    result['summary']['passed'] = int(value)
                elif 'failed' in key:
                    result['summary']['failed'] = int(value)
                elif 'skipped' in key:
                    result['summary']['not_executed'] = int(value)
                elif 'total tests' in key or key == 'total tests':
                    result['summary']['total'] = int(value)
                elif 'duration' in key:
                    # Parse duration HH:MM:SS
                    result['run_info']['total_duration_seconds'] = ExtentSummaryParser._parse_duration(value)
            
            return result
            
        except Exception as e:
            print(f"Error parsing Extent summary: {e}")
            return None
    
    @staticmethod
    def _parse_duration(duration_str: str) -> float:
        """Parse duration string (HH:MM:SS) to seconds."""
        try:
            parts = duration_str.split(':')
            if len(parts) == 3:
                hours = int(parts[0])
                minutes = int(parts[1])
                seconds = int(parts[2])
                return hours * 3600 + minutes * 60 + seconds
            elif len(parts) == 2:
                minutes = int(parts[0])
                seconds = int(parts[1])
                return minutes * 60 + seconds
        except (ValueError, IndexError):
            pass
        return 0


class PostmanExcelParser:
    """Parser for Newman (Postman) Excel reports."""
    
    @staticmethod
    def parse(excel_content: bytes) -> dict:
        """
        Parse Newman Excel report to extract test results.
        
        Excel columns:
        - System Name
        - Scenario Name  
        - Description/Testcase
        - Status/Failure (Passed/Failed)
        - Execution Time
        - Failure Message
        
        Returns:
            Dictionary containing run_info, summary, and individual tests
        """
        try:
            from openpyxl import load_workbook
            
            result = {
                'run_info': {
                    'id': '',
                    'name': 'Postman Collection Report',
                    'total_duration_seconds': 0,
                },
                'summary': {
                    'total': 0,
                    'passed': 0,
                    'failed': 0,
                    'not_executed': 0,
                },
                'tests': []
            }
            
            # Load workbook from bytes
            wb = load_workbook(filename=io.BytesIO(excel_content), read_only=True)
            ws = wb.active
            
            # Get headers from first row
            headers = []
            for cell in next(ws.iter_rows(min_row=1, max_row=1)):
                headers.append(str(cell.value or '').lower().strip())
            
            # Find column indices - order matters for disambiguation
            col_map = {}
            for i, h in enumerate(headers):
                if 'system' in h and 'name' in h:
                    col_map['system'] = i
                elif 'scenario' in h and 'name' in h:
                    col_map['scenario'] = i
                elif 'description' in h or 'testcase' in h:
                    col_map['testcase'] = i
                elif 'status' in h:  # "status/failure" has "status"
                    col_map['status'] = i
                elif 'execution' in h and 'time' in h:
                    col_map['duration'] = i
                elif 'failure' in h and 'message' in h:  # "failure message"
                    col_map['message'] = i
            
            total_duration = 0.0
            
            # Process data rows
            for row in ws.iter_rows(min_row=2):
                row_values = [cell.value for cell in row]
                
                # Skip empty rows
                if not any(row_values):
                    continue
                
                # Extract values
                system_name = str(row_values[col_map.get('system', 0)] or '')
                scenario_name = str(row_values[col_map.get('scenario', 1)] or '')
                testcase = str(row_values[col_map.get('testcase', 2)] or '')
                status = str(row_values[col_map.get('status', 3)] or '').strip()
                duration_val = row_values[col_map.get('duration', 4)]
                failure_msg = str(row_values[col_map.get('message', 5)] or '') if col_map.get('message') else None
                
                # Parse duration (in seconds)
                duration_ms = 0.0
                if duration_val:
                    try:
                        duration_ms = float(duration_val) * 1000  # Convert to ms
                        total_duration += float(duration_val)
                    except (ValueError, TypeError):
                        pass
                
                # Normalize status
                outcome = 'Passed' if status.lower() == 'passed' else 'Failed'
                
                # Build test class from system name + scenario name
                test_class = f"{system_name}.{scenario_name}" if system_name else scenario_name
                
                test_data = {
                    'test_id': '',
                    'execution_id': '',
                    'test_name': testcase[:1000] if testcase else scenario_name[:1000],
                    'test_class': test_class[:1000] if test_class else None,
                    'computer_name': None,
                    'outcome': outcome,
                    'duration_ms': duration_ms,
                    'start_time': None,
                    'end_time': None,
                    'error_message': failure_msg[:2000] if failure_msg else None,
                    'stack_trace': None,
                    'stdout': None,
                }
                
                result['tests'].append(test_data)
                
                # Update summary
                result['summary']['total'] += 1
                if outcome == 'Passed':
                    result['summary']['passed'] += 1
                else:
                    result['summary']['failed'] += 1
            
            result['run_info']['total_duration_seconds'] = total_duration
            
            wb.close()
            return result
            
        except ImportError:
            print("Error: openpyxl not installed. Run: pip install openpyxl")
            return None
        except Exception as e:
            print(f"Error parsing Postman Excel: {e}")
            return None


class GitHubTestsExtractor:
    """
    Extracts test results from GitHub Actions workflows.
    
    Supports multiple repositories and report formats.
    Downloads artifacts, parses them with the appropriate parser,
    and stores results in the database.
    """
    
    def __init__(self, db_connection, token=None, configs=None):
        """
        Initialize the extractor.
        
        Args:
            db_connection: Database connection object with engine attribute
            token: GitHub token (defaults to GITHUB_TOKEN env var)
            configs: List of repository configurations (defaults to GITHUB_TEST_CONFIGS)
        """
        self.db = db_connection
        self.token = token or os.getenv('GITHUB_TOKEN')
        
        # Parse configurations
        if configs:
            self.configs = configs if isinstance(configs, list) else [configs]
        else:
            self.configs = self._load_configs_from_env()
        
        self.headers = {
            'Accept': 'application/vnd.github+json',
            'Authorization': f'Bearer {self.token}',
            'X-GitHub-Api-Version': '2022-11-28'
        }
        
        self.base_url = 'https://api.github.com'
        
        # Parser map
        self.parsers = {
            'trx': self._parse_trx_artifact,
            'extent_summary': self._parse_extent_artifact,
            'postman_html': self._parse_postman_artifact,
        }
    
    def _load_configs_from_env(self) -> list:
        """Load repository configurations from environment variables."""
        # Try new JSON format first
        configs_json = os.getenv('GITHUB_TEST_CONFIGS')
        if configs_json:
            try:
                return json.loads(configs_json)
            except json.JSONDecodeError as e:
                print(f"Error parsing GITHUB_TEST_CONFIGS JSON: {e}")
        
        # Fall back to legacy format
        owner = os.getenv('GITHUB_TEST_OWNER', 'Pathlock')
        repo = os.getenv('GITHUB_TEST_REPO', 'pathlock-plc')
        branch = os.getenv('GITHUB_TEST_BRANCH', 'cloud/dev')
        workflows_env = os.getenv('GITHUB_TEST_WORKFLOWS', '')
        patterns_env = os.getenv('GITHUB_TEST_ARTIFACT_PATTERNS', '')
        
        if workflows_env:
            workflows = [w.strip() for w in workflows_env.split(',') if w.strip()]
        else:
            workflows = ['Nightly Full Unit Test Run', 'Nightly Full Integration Test Run']
        
        if patterns_env:
            patterns = [p.strip() for p in patterns_env.split(',') if p.strip()]
        else:
            patterns = ['test-results', 'nightly-test-results']
        
        return [{
            'owner': owner,
            'repo': repo,
            'branch': branch,
            'workflows': workflows,
            'artifact_patterns': patterns,
            'report_format': 'trx',
            'team': 'default'
        }]
    
    def setup_tables(self):
        """Create the database tables if they don't exist."""
        try:
            with self.db.engine.connect() as connection:
                # Create github_test_runs table with team column
                connection.execute(text("""
                    CREATE TABLE IF NOT EXISTS github_test_runs (
                        id SERIAL PRIMARY KEY,
                        run_id BIGINT NOT NULL,
                        workflow_name VARCHAR(500),
                        repository VARCHAR(200),
                        branch VARCHAR(200),
                        commit_sha VARCHAR(40),
                        run_number INTEGER,
                        run_started_at TIMESTAMP,
                        run_completed_at TIMESTAMP,
                        conclusion VARCHAR(50),
                        test_run_id VARCHAR(100),
                        test_run_name VARCHAR(500),
                        computer_name VARCHAR(200),
                        total_tests INTEGER DEFAULT 0,
                        passed_tests INTEGER DEFAULT 0,
                        failed_tests INTEGER DEFAULT 0,
                        skipped_tests INTEGER DEFAULT 0,
                        total_duration_seconds FLOAT DEFAULT 0,
                        artifact_name VARCHAR(200),
                        team VARCHAR(100),
                        created_date TIMESTAMP DEFAULT CURRENT_TIMESTAMP
                    )
                """))
                
                # Add team column if it doesn't exist (for existing installations)
                connection.execute(text("""
                    DO $$
                    BEGIN
                        IF NOT EXISTS (
                            SELECT 1 FROM information_schema.columns 
                            WHERE table_name = 'github_test_runs' AND column_name = 'team'
                        ) THEN
                            ALTER TABLE github_test_runs ADD COLUMN team VARCHAR(100);
                        END IF;
                    END $$;
                """))
                
                # Create unique constraint on run_id + repository (allowing same run_id across repos)
                connection.execute(text("""
                    DO $$
                    BEGIN
                        IF NOT EXISTS (
                            SELECT 1 FROM pg_constraint 
                            WHERE conname = 'github_test_runs_run_id_repo_unique'
                        ) THEN
                            -- First drop the old unique constraint on run_id if it exists
                            ALTER TABLE github_test_runs DROP CONSTRAINT IF EXISTS github_test_runs_run_id_key;
                            -- Create new composite unique constraint
                            ALTER TABLE github_test_runs 
                            ADD CONSTRAINT github_test_runs_run_id_repo_unique UNIQUE (run_id, repository);
                        END IF;
                    END $$;
                """))
                
                # Create github_test_results table
                connection.execute(text("""
                    CREATE TABLE IF NOT EXISTS github_test_results (
                        id SERIAL PRIMARY KEY,
                        run_id BIGINT NOT NULL,
                        test_id VARCHAR(100),
                        execution_id VARCHAR(100),
                        test_name VARCHAR(1000) NOT NULL,
                        test_class VARCHAR(1000),
                        outcome VARCHAR(50) NOT NULL,
                        duration_ms FLOAT,
                        start_time TIMESTAMP,
                        end_time TIMESTAMP,
                        error_message TEXT,
                        stack_trace TEXT,
                        stdout TEXT,
                        created_date TIMESTAMP DEFAULT CURRENT_TIMESTAMP
                    )
                """))
                
                # Create indexes
                connection.execute(text("""
                    CREATE INDEX IF NOT EXISTS idx_test_runs_workflow 
                    ON github_test_runs(workflow_name, run_started_at)
                """))
                connection.execute(text("""
                    CREATE INDEX IF NOT EXISTS idx_test_runs_branch 
                    ON github_test_runs(branch, run_started_at)
                """))
                connection.execute(text("""
                    CREATE INDEX IF NOT EXISTS idx_test_runs_team 
                    ON github_test_runs(team, run_started_at)
                """))
                connection.execute(text("""
                    CREATE INDEX IF NOT EXISTS idx_test_runs_repo 
                    ON github_test_runs(repository, run_started_at)
                """))
                connection.execute(text("""
                    CREATE INDEX IF NOT EXISTS idx_test_results_run_id 
                    ON github_test_results(run_id)
                """))
                connection.execute(text("""
                    CREATE INDEX IF NOT EXISTS idx_test_results_outcome 
                    ON github_test_results(outcome)
                """))
                connection.execute(text("""
                    CREATE INDEX IF NOT EXISTS idx_test_results_name 
                    ON github_test_results(test_name)
                """))
                
                connection.commit()
                print("------>GitHub test tables created/verified")
                
        except SQLAlchemyError as e:
            print(f"Error setting up GitHub test tables: {e}")
            raise
    
    def is_run_synced(self, run_id: int, repository: str) -> bool:
        """Check if a workflow run has already been synced for a specific repo."""
        try:
            with self.db.engine.connect() as connection:
                result = connection.execute(
                    text("SELECT 1 FROM github_test_runs WHERE run_id = :run_id AND repository = :repo"),
                    {"run_id": run_id, "repo": repository}
                ).first()
                return result is not None
        except SQLAlchemyError:
            return False
    
    def fetch_workflow_runs(self, config: dict, since_date: datetime = None, limit: int = 100) -> list:
        """
        Fetch workflow runs from GitHub API for a specific repository configuration.
        
        Args:
            config: Repository configuration dictionary
            since_date: Only fetch runs after this date
            limit: Maximum number of runs to fetch per workflow
            
        Returns:
            List of workflow run dictionaries
        """
        owner = config['owner']
        repo = config['repo']
        branch = config.get('branch')
        workflows = config.get('workflows', [])
        
        all_runs = []
        
        # First, get workflow IDs for our target workflows
        workflows_url = f'{self.base_url}/repos/{owner}/{repo}/actions/workflows'
        response = requests.get(workflows_url, headers=self.headers)
        
        if response.status_code != 200:
            print(f"Error fetching workflows for {owner}/{repo}: {response.status_code} - {response.text}")
            return []
        
        workflow_map = {}
        for wf in response.json().get('workflows', []):
            if wf['name'] in workflows:
                workflow_map[wf['name']] = wf['id']
        
        # Fetch runs for each workflow separately
        for workflow_name, workflow_id in workflow_map.items():
            workflow_runs = []
            page = 1
            
            while len(workflow_runs) < limit:
                params = {
                    'status': 'completed',
                    'per_page': min(limit - len(workflow_runs), 100),
                    'page': page
                }
                
                # Only filter by branch if specified
                if branch:
                    params['branch'] = branch
                
                url = f'{self.base_url}/repos/{owner}/{repo}/actions/workflows/{workflow_id}/runs'
                response = requests.get(url, headers=self.headers, params=params)
                
                if response.status_code != 200:
                    print(f"Error fetching runs for {workflow_name}: {response.status_code}")
                    break
                
                runs = response.json().get('workflow_runs', [])
                if not runs:
                    break  # No more runs
                
                for run in runs:
                    # Filter by date if specified
                    if since_date:
                        run_date = datetime.fromisoformat(run['created_at'].replace('Z', '+00:00'))
                        if run_date.replace(tzinfo=None) < since_date:
                            # Runs are returned in reverse chronological order
                            # If we hit a run before since_date, stop pagination
                            break
                    
                    # Add config reference to run for later use
                    run['_config'] = config
                    workflow_runs.append(run)
                    
                    if len(workflow_runs) >= limit:
                        break
                else:
                    # Continue pagination if we didn't break
                    page += 1
                    continue
                break  # Break outer loop if inner loop broke
            
            all_runs.extend(workflow_runs)
        
        # Sort all runs by date (newest first)
        all_runs.sort(key=lambda r: r['created_at'], reverse=True)
        
        return all_runs
    
    def fetch_artifacts(self, run_id: int, owner: str, repo: str, 
                       workflow_name: str = None, artifact_patterns: list = None) -> list:
        """
        Fetch artifacts for a specific workflow run.
        """
        url = f'{self.base_url}/repos/{owner}/{repo}/actions/runs/{run_id}/artifacts'
        response = requests.get(url, headers=self.headers)
        
        if response.status_code != 200:
            print(f"Error fetching artifacts for run {run_id}: {response.status_code}")
            return []
        
        artifacts = response.json().get('artifacts', [])
        
        # Filter out expired artifacts
        valid_artifacts = [a for a in artifacts if not a['expired']]
        
        # If no patterns specified, return all valid artifacts
        if not artifact_patterns:
            return valid_artifacts
        
        # Define preferred artifact names per workflow type (for TRX format)
        preferred_artifacts = {
            'unit': ['nightly-test-results-all-tests', 'nightly-test-results'],
            'integration': ['test-results', 'nightly-integration-test-results-summary'],
        }
        
        # Determine workflow type for TRX preference logic
        workflow_type = None
        if workflow_name:
            workflow_lower = workflow_name.lower()
            if 'integration' in workflow_lower:
                workflow_type = 'integration'
            elif 'unit' in workflow_lower:
                workflow_type = 'unit'
        
        # Try to find preferred artifact first (for TRX workflows)
        if workflow_type and workflow_type in preferred_artifacts:
            for preferred_name in preferred_artifacts[workflow_type]:
                for artifact in valid_artifacts:
                    if artifact['name'].lower() == preferred_name.lower():
                        return [artifact]  # Return only the preferred one
        
        # Filter by patterns
        matching_artifacts = []
        for artifact in valid_artifacts:
            artifact_name = artifact['name'].lower()
            for pattern in artifact_patterns:
                if pattern.lower() in artifact_name or artifact_name == pattern.lower():
                    matching_artifacts.append(artifact)
                    break
        
        # Return first match only to avoid duplicates
        return matching_artifacts[:1] if matching_artifacts else []
    
    def download_artifact(self, artifact_id: int, owner: str, repo: str) -> bytes:
        """Download an artifact and return its content."""
        url = f'{self.base_url}/repos/{owner}/{repo}/actions/artifacts/{artifact_id}/zip'
        response = requests.get(url, headers=self.headers, allow_redirects=True)
        
        if response.status_code != 200:
            print(f"Error downloading artifact {artifact_id}: {response.status_code}")
            return None
        
        return response.content
    
    def _parse_trx_artifact(self, zip_content: bytes) -> dict:
        """Extract and parse TRX files from a ZIP archive."""
        try:
            with zipfile.ZipFile(io.BytesIO(zip_content), 'r') as zf:
                for filename in zf.namelist():
                    if filename.endswith('.trx'):
                        content = zf.read(filename).decode('utf-8')
                        result = TRXParser.parse(content)
                        if result:
                            result['_source_file'] = filename
                            return result
        except zipfile.BadZipFile as e:
            print(f"Error extracting TRX ZIP: {e}")
        return None
    
    def _parse_extent_artifact(self, zip_content: bytes) -> dict:
        """Extract and parse Extent summary from a ZIP archive."""
        try:
            with zipfile.ZipFile(io.BytesIO(zip_content), 'r') as zf:
                for filename in zf.namelist():
                    if filename.endswith('test-summary.txt'):
                        content = zf.read(filename).decode('utf-8')
                        result = ExtentSummaryParser.parse(content)
                        if result:
                            result['_source_file'] = filename
                            return result
        except zipfile.BadZipFile as e:
            print(f"Error extracting Extent ZIP: {e}")
        return None
    
    def _parse_postman_artifact(self, zip_content: bytes) -> dict:
        """Extract and parse Postman Excel report from a ZIP archive."""
        try:
            with zipfile.ZipFile(io.BytesIO(zip_content), 'r') as zf:
                # Look for Excel files first (preferred)
                for filename in zf.namelist():
                    if filename.endswith('.xlsx') and 'report' in filename.lower():
                        content = zf.read(filename)
                        result = PostmanExcelParser.parse(content)
                        if result:
                            result['_source_file'] = filename
                            return result
                
                # Fallback: try any Excel file
                for filename in zf.namelist():
                    if filename.endswith('.xlsx'):
                        content = zf.read(filename)
                        result = PostmanExcelParser.parse(content)
                        if result:
                            result['_source_file'] = filename
                            return result
        except zipfile.BadZipFile as e:
            print(f"Error extracting Postman ZIP: {e}")
        return None
    
    def store_test_run(self, run_data: dict, parsed_result: dict, artifact_name: str, 
                      config: dict) -> bool:
        """Store a test run and its results in the database."""
        try:
            with self.db.engine.connect() as connection:
                run_info = parsed_result['run_info']
                summary = parsed_result['summary']
                
                repository = f"{config['owner']}/{config['repo']}"
                team = config.get('team', 'default')
                
                # Get computer name from first test result if available
                computer_name = None
                if parsed_result.get('tests'):
                    computer_name = parsed_result['tests'][0].get('computer_name')
                
                # Insert test run with ON CONFLICT for run_id + repository
                connection.execute(text("""
                    INSERT INTO github_test_runs (
                        run_id, workflow_name, repository, branch, commit_sha,
                        run_number, run_started_at, run_completed_at, conclusion,
                        test_run_id, test_run_name, computer_name,
                        total_tests, passed_tests, failed_tests, skipped_tests,
                        total_duration_seconds, artifact_name, team
                    ) VALUES (
                        :run_id, :workflow_name, :repository, :branch, :commit_sha,
                        :run_number, :run_started_at, :run_completed_at, :conclusion,
                        :test_run_id, :test_run_name, :computer_name,
                        :total_tests, :passed_tests, :failed_tests, :skipped_tests,
                        :total_duration_seconds, :artifact_name, :team
                    )
                    ON CONFLICT (run_id, repository) DO UPDATE SET
                        total_tests = EXCLUDED.total_tests,
                        passed_tests = EXCLUDED.passed_tests,
                        failed_tests = EXCLUDED.failed_tests,
                        skipped_tests = EXCLUDED.skipped_tests,
                        total_duration_seconds = EXCLUDED.total_duration_seconds,
                        team = EXCLUDED.team
                """), {
                    'run_id': run_data['id'],
                    'workflow_name': run_data.get('name', ''),
                    'repository': repository,
                    'branch': run_data.get('head_branch', config.get('branch', '')),
                    'commit_sha': run_data.get('head_sha', '')[:40],
                    'run_number': run_data.get('run_number'),
                    'run_started_at': datetime.fromisoformat(run_data['created_at'].replace('Z', '+00:00')).replace(tzinfo=None) if run_data.get('created_at') else None,
                    'run_completed_at': datetime.fromisoformat(run_data['updated_at'].replace('Z', '+00:00')).replace(tzinfo=None) if run_data.get('updated_at') else None,
                    'conclusion': run_data.get('conclusion', ''),
                    'test_run_id': run_info.get('id', ''),
                    'test_run_name': run_info.get('name', '')[:500],
                    'computer_name': computer_name,
                    'total_tests': summary['total'],
                    'passed_tests': summary['passed'],
                    'failed_tests': summary['failed'],
                    'skipped_tests': summary['not_executed'],
                    'total_duration_seconds': run_info.get('total_duration_seconds', 0),
                    'artifact_name': artifact_name,
                    'team': team
                })
                
                # Delete existing test results for this run (in case of re-sync)
                connection.execute(text("""
                    DELETE FROM github_test_results WHERE run_id = :run_id
                """), {'run_id': run_data['id']})
                
                # Insert individual test results (only for formats that provide them)
                for test in parsed_result.get('tests', []):
                    connection.execute(text("""
                        INSERT INTO github_test_results (
                            run_id, test_id, execution_id, test_name, test_class,
                            outcome, duration_ms, start_time, end_time,
                            error_message, stack_trace, stdout
                        ) VALUES (
                            :run_id, :test_id, :execution_id, :test_name, :test_class,
                            :outcome, :duration_ms, :start_time, :end_time,
                            :error_message, :stack_trace, :stdout
                        )
                    """), {
                        'run_id': run_data['id'],
                        'test_id': test.get('test_id', '')[:100],
                        'execution_id': test.get('execution_id', '')[:100],
                        'test_name': test.get('test_name', '')[:1000],
                        'test_class': test.get('test_class', '')[:1000] if test.get('test_class') else None,
                        'outcome': test.get('outcome', ''),
                        'duration_ms': test.get('duration_ms'),
                        'start_time': test.get('start_time'),
                        'end_time': test.get('end_time'),
                        'error_message': test.get('error_message'),
                        'stack_trace': test.get('stack_trace'),
                        'stdout': test.get('stdout')
                    })
                
                connection.commit()
                return True
                
        except SQLAlchemyError as e:
            print(f"Error storing test run {run_data['id']}: {e}")
            return False
    
    def sync_test_runs(self, days_back: int = None, initial_sync: bool = False) -> int:
        """
        Sync test runs from GitHub Actions for all configured repositories.
        
        Args:
            days_back: Number of days to look back (default: 1 for daily, 60 for initial)
            initial_sync: If True, fetch last 2 months of data
            
        Returns:
            Number of runs processed
        """
        # Ensure tables exist
        self.setup_tables()
        
        if not self.token:
            print("------>Skipping GitHub tests sync (GITHUB_TOKEN not configured)")
            return 0
        
        # Determine how far back to sync
        if initial_sync:
            since_date = datetime.now() - timedelta(days=60)  # 2 months
            print(f"------>Initial sync: fetching test runs from last 60 days")
        elif days_back:
            since_date = datetime.now() - timedelta(days=days_back)
            print(f"------>Fetching test runs from last {days_back} day(s)")
        else:
            # Default: 1 day for daily sync
            since_date = datetime.now() - timedelta(days=1)
            print(f"------>Fetching test runs from last 1 day")
        
        total_processed = 0
        
        # Process each repository configuration
        for config in self.configs:
            owner = config['owner']
            repo = config['repo']
            team = config.get('team', 'default')
            report_format = config.get('report_format', 'trx')
            workflows = config.get('workflows', [])
            
            print(f"\n------>Processing {owner}/{repo} (team: {team})")
            print(f"        Workflows: {', '.join(workflows)}")
            print(f"        Report format: {report_format}")
            
            # Fetch workflow runs for this config
            runs = self.fetch_workflow_runs(config, since_date=since_date, 
                                           limit=200 if initial_sync else 50)
            print(f"        Found {len(runs)} workflow runs to process")
            
            for run in runs:
                run_id = run['id']
                repository = f"{owner}/{repo}"
                
                # Skip if already synced
                if self.is_run_synced(run_id, repository):
                    continue
                
                workflow_name = run['name']
                print(f"        Processing run #{run['run_number']}: {workflow_name} (ID: {run_id})")
                
                # Fetch artifacts
                artifacts = self.fetch_artifacts(
                    run_id, owner, repo,
                    workflow_name=workflow_name,
                    artifact_patterns=config.get('artifact_patterns')
                )
                
                if not artifacts:
                    print(f"          No matching test artifacts found")
                    continue
                
                # Process the first matching artifact
                artifact = artifacts[0]
                artifact_name = artifact['name']
                print(f"          Downloading artifact: {artifact_name}")
                
                # Download artifact
                zip_content = self.download_artifact(artifact['id'], owner, repo)
                if not zip_content:
                    continue
                
                # Parse with appropriate parser
                parser_func = self.parsers.get(report_format)
                if not parser_func:
                    print(f"          Unknown report format: {report_format}")
                    continue
                
                parsed = parser_func(zip_content)
                if not parsed:
                    print(f"          Failed to parse artifact")
                    continue
                
                summary = parsed['summary']
                print(f"          Tests: {summary['total']} total, {summary['passed']} passed, {summary['failed']} failed, {summary['not_executed']} skipped")
                
                # Store results
                if self.store_test_run(run, parsed, artifact_name, config):
                    total_processed += 1
                    print(f"          ✓ Stored test results")
        
        print(f"\n------>Processed {total_processed} test runs total")
        return total_processed
    
    def get_sync_summary(self) -> dict:
        """Get a summary of synced test data."""
        try:
            with self.db.engine.connect() as connection:
                total_runs = connection.execute(text(
                    "SELECT COUNT(*) FROM github_test_runs"
                )).scalar()
                
                total_tests = connection.execute(text(
                    "SELECT COUNT(*) FROM github_test_results"
                )).scalar()
                
                latest_run = connection.execute(text(
                    "SELECT MAX(run_started_at) FROM github_test_runs"
                )).scalar()
                
                # Get counts by team
                team_counts = connection.execute(text("""
                    SELECT team, COUNT(*) as count 
                    FROM github_test_runs 
                    GROUP BY team
                """)).fetchall()
                
                return {
                    'total_runs': total_runs,
                    'total_tests': total_tests,
                    'latest_run': latest_run,
                    'by_team': {row[0]: row[1] for row in team_counts}
                }
        except SQLAlchemyError:
            return {'total_runs': 0, 'total_tests': 0, 'latest_run': None, 'by_team': {}}


def main():
    """Standalone execution for testing."""
    import argparse
    
    parser = argparse.ArgumentParser(description='Sync GitHub Actions test results')
    parser.add_argument('--days', type=int, default=1, help='Number of days to sync (default: 1)')
    parser.add_argument('--initial', action='store_true', help='Run initial sync (2 months)')
    args = parser.parse_args()
    
    # Create database connection
    from export_ado import get_database_connection
    db = get_database_connection()
    
    # Create extractor (will use GITHUB_TEST_CONFIGS from env)
    extractor = GitHubTestsExtractor(db_connection=db)
    
    print(f"Loaded {len(extractor.configs)} repository configuration(s)")
    for i, config in enumerate(extractor.configs):
        print(f"  {i+1}. {config['owner']}/{config['repo']} - team: {config.get('team', 'default')}")
    
    # Run sync
    if args.initial:
        processed = extractor.sync_test_runs(initial_sync=True)
    else:
        processed = extractor.sync_test_runs(days_back=args.days)
    
    # Print summary
    summary = extractor.get_sync_summary()
    print(f"\n=== Sync Complete ===")
    print(f"Total runs in DB: {summary['total_runs']}")
    print(f"Total tests in DB: {summary['total_tests']}")
    print(f"Latest run: {summary['latest_run']}")
    print(f"By team: {summary.get('by_team', {})}")


if __name__ == '__main__':
    main()
